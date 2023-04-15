import json

import openpyxl as openpyxl
import requests
from bs4 import BeautifulSoup
import pymongo


# AllMobileDetails = dict()
# mobileIndex = 0
# for link in mobileList:
#
#     for detail in details:
#
#         if str(detail) in AllMobileDetails:
#             AllMobileDetails[str(detail)][mobileIndex] = str(details[detail])
#         else:
#             AllMobileDetails[str(detail)] = {}
#             AllMobileDetails[str(detail)][mobileIndex] = str(details[detail])
#
#     mobileIndex = mobileIndex + 1
#
#     # wb.save(str(details["MobileName"]) + ".xlsx")
#     # jsonConverted = json.dumps(details)


# def details_to_excel(all_details):
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#     index = 1
#     for mobile_detail in all_details:
#
#         sheet.cell(column=index, row=1).value = mobile_detail
#         for Mobile in AllMobileDetails[mobile_detail]:
#             sheet.cell(column=index, row=Mobile + 2).value = str(AllMobileDetails[mobile_detail][Mobile])
#         index = index + 1
#
#     wb.save(str("all") + ".xlsx")


def upload_to_mongodb(title, details):
    myClient = pymongo.MongoClient("mongodb+srv://detailmobile65:L2PcRKlQA2UJFDXd@cluster0.qsgcfl3.mongodb.net/")
    myDb = myClient["MobileDetail"]
    myCol = myDb[title]
    x = myCol.insert_one(details)
    print(x.inserted_id)


def links_from_company_what_mobile(path):
    mainSoupLink = BeautifulSoup(requests.get(path).content, "html.parser")
    mobileLink = mainSoupLink.find('div', {'class': 'mobiles'})
    mobileList = mobileLink.find_all("a", {'class': 'BiggerText'}, href=True)
    return map(lambda x: x["href"], mobileList)


def link_to_map_what_mobile(path):
    soupLink = BeautifulSoup(requests.get("https://www.whatmobile.com.pk/" + path).content, "html.parser")
    details = {}
    mobileName = soupLink.find('h2', {'class': 'Heading1'})

    if mobileName.text == " Top Mobile Phones ":
        return None

    details["MobileName"] = mobileName.text.replace("detailed specifications", "")

    columns = soupLink.findAll('tr')
    for column in columns:
        try:
            td = ""
            th = str(column.find("th").text).replace("\n", "").replace("\xa0", "").replace(" ", "")

            if len(column.findAll("td")) > 1:
                td = str(column.findAll("td")[1].text).replace("\n", "").replace("\xa0", "")
            else:
                td = str(column.findAll("td")[0].text).replace("\n", "").replace("\xa0", "")

            if th == "":
                details["Extra"] = details["Extra"] + " , " + td
            else:
                details[th] = td
        except:
            print("Error : " + str(column.find("th")))
    try:
        if "N/A" not in details["Weight"]:
            details["Weight"] = int(details["Weight"].replace(" ", "").replace("g", ""))

        if "N/A" not in details["Size"]:
            details["Size"] = float(details["Size"].replace("Inches", "").replace(" ", ""))

        if "N/A" not in details["Price"] and "Discontinued" not in details["Price"]:

            Price = details["Price"].replace("Price in Rs: ", "").replace("Price in USD: ", "").replace(",", "")

            v = Price.split(" ")

            if len(v) == 2:
                details["PriceRs"] = int(v[0])
                details["PriceUSD"] = int(v[1].replace("$", ""))

        if "Yes" in details["Card"]:
            details["Card"] = 1
        else:
            details["Card"] = 0

        if "Yes" in details["NFC"]:
            details["NFC"] = 1
        else:
            details["NFC"] = 0

        if "Yes" in details["Torch"]:
            details["Torch"] = 1
        else:
            details["Torch"] = 0

        details.pop('Ratings', None)
    except:
        return None

    return details


def main():
    print("Hello World!")

    main_link = "https://www.whatmobile.com.pk/Vivo_Mobiles_Prices"
    company_name = "Vivo"
    all_mobile_list = dict()

    mobile_link_list = links_from_company_what_mobile(main_link)
    for mobile_link in mobile_link_list:
        print(mobile_link)
        mobile_detail = link_to_map_what_mobile(mobile_link)
        if mobile_detail is None:
            print("Error Fuck you")
            continue
        all_mobile_list[str(mobile_detail["MobileName"])] = mobile_detail
        upload_to_mongodb(company_name, mobile_detail)

    print("Code End")


if __name__ == "__main__":
    main()
